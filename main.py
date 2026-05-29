"""倉庫在庫管理アプリ — FastAPI 本体。

工場倉庫の原料の入庫・出庫を記録し、現在庫を自動算出する社内 Web アプリ。
このファイルを直接実行すると 0.0.0.0:8000 でサーバーが起動し、
社内の各 PC からブラウザで http://サーバー名:8000/ にアクセスできる。
"""

import datetime
import hashlib
import io
import os
import re
import socket
import sqlite3
import sys
from contextlib import asynccontextmanager
from typing import Literal, Optional

import openpyxl
import xlrd
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.responses import FileResponse, Response
from pydantic import BaseModel, ConfigDict, Field

from database import get_conn, init_db


def resource_dir() -> str:
    """同梱リソース(static)の置き場所。exe では展開先、開発時はスクリプト位置。"""
    if getattr(sys, "frozen", False):
        return sys._MEIPASS  # type: ignore[attr-defined]
    return os.path.dirname(os.path.abspath(__file__))


STATIC_DIR = os.path.join(resource_dir(), "static")
PORT = 8000


@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    yield


app = FastAPI(title="倉庫在庫管理", lifespan=lifespan)


# --- 入力モデル ---------------------------------------------------------------

class MaterialIn(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)

    code: str = Field(min_length=1, max_length=50)
    name: str = Field(min_length=1, max_length=100)
    product_name: str = Field(default="", max_length=100)
    pack_size: str = Field(default="", max_length=50)
    unit: str = Field(default="個", max_length=20)
    reorder_point: float = Field(default=0, ge=0)
    supplier: str = Field(default="", max_length=100)
    location: str = Field(default="倉庫", max_length=50)


class TransactionIn(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)

    material_id: int
    type: Literal["in", "out"]
    quantity: float = Field(gt=0)
    line: str = Field(default="", max_length=100)
    note: str = Field(default="", max_length=300)


class BatchOutItem(BaseModel):
    material_id: int
    quantity: float = Field(gt=0)


class BatchOutIn(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)

    line: str = Field(default="", max_length=100)
    note: str = Field(default="", max_length=300)
    items: list[BatchOutItem] = Field(default_factory=list)


class RegisterIn(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)

    location: str = Field(default="倉庫", max_length=50)
    ids: list[int] = Field(default_factory=list)


class LocationIn(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)

    name: str = Field(min_length=1, max_length=50)


class OutboundLineIn(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)

    name: str = Field(min_length=1, max_length=100)


class OutboundLineUpdate(BaseModel):
    """出庫先ラインの名前と品目（material_id の並び）をまとめて更新する。"""

    model_config = ConfigDict(str_strip_whitespace=True)

    name: str = Field(min_length=1, max_length=100)
    items: list[int] = Field(default_factory=list)


# --- 在庫計算ヘルパー ---------------------------------------------------------

def _stock(conn: sqlite3.Connection, material_id: int) -> float:
    """現在庫。入庫は納品日が今日以前のものだけを数える（未来＝入荷予定は除外）。"""
    row = conn.execute(
        "SELECT COALESCE(SUM(CASE "
        "WHEN type = 'in' AND date(created_at) <= date('now', 'localtime') "
        "THEN quantity "
        "WHEN type = 'out' THEN -quantity ELSE 0 END), 0) AS s "
        "FROM transactions WHERE material_id = ?",
        (material_id,),
    ).fetchone()
    return row["s"]


def _cell_str(v) -> str:
    """Excel セルの値を文字列に整える（10000.0 → "10000" など）。"""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _norm(s: str) -> str:
    """見出しセルの改行・空白を取り除いて比較しやすくする。"""
    return s.replace("\n", "").replace("\r", "").replace(" ", "").replace("　", "")


def _parse_inbound_date(v) -> str:
    """納品日を YYYY-MM-DD 文字列にする。年が無ければ今年とみなす。"""
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.date):
        return v.strftime("%Y-%m-%d")
    m = re.match(r"(?:(\d{2,4})/)?(\d{1,2})/(\d{1,2})$", _cell_str(v))
    if m:
        year = datetime.date.today().year if m.group(1) is None else int(m.group(1))
        if year < 100:
            year += 2000
        try:
            return datetime.date(
                year, int(m.group(2)), int(m.group(3))
            ).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return datetime.date.today().strftime("%Y-%m-%d")


def _row_fingerprint(values) -> str:
    """入庫記録1行の指紋（再取り込み時の二重登録を防ぐ）。"""
    return hashlib.sha1("|".join(values).encode("utf-8")).hexdigest()[:16]


def _read_xlsx(content: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        return [list(r) for r in wb.active.iter_rows(values_only=True)]
    finally:
        wb.close()


def _read_xls(content: bytes):
    book = xlrd.open_workbook(file_contents=content)
    sheet = book.sheet_by_index(0)
    rows = []
    for r in range(sheet.nrows):
        row = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    row.append(xlrd.xldate_as_datetime(cell.value, book.datemode))
                except Exception:
                    row.append(cell.value)
            else:
                row.append(cell.value)
        rows.append(row)
    return rows


def _read_workbook(filename: str, content: bytes):
    """アップロードされた Excel(.xlsx / .xls) を行のリストとして読み込む。"""
    is_xls = (filename or "").lower().endswith(".xls")
    readers = [_read_xls, _read_xlsx] if is_xls else [_read_xlsx, _read_xls]
    for reader in readers:
        try:
            return reader(content)
        except Exception:
            continue
    raise HTTPException(
        400,
        "Excel ファイルとして読み込めませんでした。.xlsx または .xls 形式か確認してください",
    )


# --- 原料マスター API ---------------------------------------------------------

@app.get("/api/materials")
def list_materials():
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT m.*,
                   COALESCE((SELECT SUM(CASE
                       WHEN type = 'in'
                            AND date(created_at) <= date('now', 'localtime')
                            THEN quantity
                       WHEN type = 'out' THEN -quantity ELSE 0 END)
                     FROM transactions t WHERE t.material_id = m.id), 0) AS stock
            FROM materials m
            WHERE m.active = 1
            ORDER BY m.sort_order, m.code
            """
        ).fetchall()
        plans = conn.execute(
            "SELECT material_id, date(created_at) AS d, quantity FROM transactions "
            "WHERE type = 'in' AND date(created_at) > date('now', 'localtime') "
            "ORDER BY created_at"
        ).fetchall()
    incoming = {}
    for p in plans:
        incoming.setdefault(p["material_id"], []).append(
            {"date": p["d"], "quantity": p["quantity"]}
        )
    result = []
    for r in rows:
        d = dict(r)
        d["low"] = d["reorder_point"] > 0 and d["stock"] <= d["reorder_point"]
        d["incoming"] = incoming.get(d["id"], [])
        result.append(d)
    return result


@app.get("/api/catalog")
def list_catalog():
    """取り込み済みだが、まだ在庫登録されていない原料（カタログ）の一覧。"""
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT id, code, name, product_name, pack_size, unit FROM materials "
            "WHERE active = 0 ORDER BY code"
        ).fetchall()
    return [dict(r) for r in rows]


@app.post("/api/materials/register")
def register_materials(reg: RegisterIn):
    """カタログから選んだ原料を、指定の保管場所で在庫登録（active=1）する。
    入力 ids の順に sort_order を割り当て、在庫一覧の並び順を保つ。"""
    if not reg.ids:
        raise HTTPException(400, "登録する原料が選ばれていません")
    with get_conn() as conn:
        cur_max = conn.execute(
            "SELECT COALESCE(MAX(sort_order), 0) AS m FROM materials"
        ).fetchone()["m"]
        registered = 0
        for i, mid in enumerate(reg.ids):
            cur = conn.execute(
                "UPDATE materials SET active = 1, location = ?, sort_order = ? "
                "WHERE id = ?",
                (reg.location, cur_max + 1 + i, mid),
            )
            registered += cur.rowcount
        return {"registered": registered}


@app.post("/api/materials/{material_id}/unregister")
def unregister_material(material_id: int):
    """在庫登録を解除し、カタログに戻す（履歴は残す）。"""
    with get_conn() as conn:
        cur = conn.execute(
            "UPDATE materials SET active = 0 WHERE id = ?", (material_id,)
        )
        if cur.rowcount == 0:
            raise HTTPException(404, "原料が見つかりません")
    return {"ok": True}


@app.post("/api/materials", status_code=201)
def create_material(m: MaterialIn):
    with get_conn() as conn:
        cur_max = conn.execute(
            "SELECT COALESCE(MAX(sort_order), 0) AS m FROM materials"
        ).fetchone()["m"]
        try:
            cur = conn.execute(
                "INSERT INTO materials "
                "(code, name, product_name, pack_size, unit, reorder_point, "
                "supplier, location, sort_order) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (m.code, m.name, m.product_name, m.pack_size, m.unit,
                 m.reorder_point, m.supplier, m.location, cur_max + 1),
            )
        except sqlite3.IntegrityError:
            raise HTTPException(400, f"原料コード「{m.code}」はすでに登録されています")
        return {"id": cur.lastrowid}


@app.put("/api/materials/{material_id}")
def update_material(material_id: int, m: MaterialIn):
    with get_conn() as conn:
        if conn.execute("SELECT 1 FROM materials WHERE id = ?", (material_id,)).fetchone() is None:
            raise HTTPException(404, "原料が見つかりません")
        try:
            conn.execute(
                "UPDATE materials SET code = ?, name = ?, product_name = ?, "
                "pack_size = ?, unit = ?, reorder_point = ?, supplier = ?, "
                "location = ? WHERE id = ?",
                (m.code, m.name, m.product_name, m.pack_size, m.unit,
                 m.reorder_point, m.supplier, m.location, material_id),
            )
        except sqlite3.IntegrityError:
            raise HTTPException(400, f"原料コード「{m.code}」はすでに登録されています")
    return {"ok": True}


@app.delete("/api/materials/{material_id}")
def delete_material(material_id: int):
    with get_conn() as conn:
        cur = conn.execute("DELETE FROM materials WHERE id = ?", (material_id,))
        if cur.rowcount == 0:
            raise HTTPException(404, "原料が見つかりません")
    return {"ok": True}


# --- 保管場所 API -------------------------------------------------------------

@app.get("/api/locations")
def list_locations():
    with get_conn() as conn:
        return [
            dict(r)
            for r in conn.execute("SELECT id, name FROM locations ORDER BY id")
        ]


@app.post("/api/locations", status_code=201)
def create_location(loc: LocationIn):
    with get_conn() as conn:
        try:
            cur = conn.execute(
                "INSERT INTO locations(name) VALUES (?)", (loc.name,)
            )
        except sqlite3.IntegrityError:
            raise HTTPException(400, f"「{loc.name}」はすでに登録されています")
        return {"id": cur.lastrowid, "name": loc.name}


@app.delete("/api/locations/{loc_id}")
def delete_location(loc_id: int):
    with get_conn() as conn:
        row = conn.execute(
            "SELECT name FROM locations WHERE id = ?", (loc_id,)
        ).fetchone()
        if row is None:
            raise HTTPException(404, "保管場所が見つかりません")
        used = conn.execute(
            "SELECT 1 FROM materials WHERE active = 1 AND location = ? LIMIT 1",
            (row["name"],),
        ).fetchone()
        if used:
            raise HTTPException(
                400, f"「{row['name']}」を使っている原料があるため削除できません"
            )
        conn.execute("DELETE FROM locations WHERE id = ?", (loc_id,))
    return {"ok": True}


# --- 出庫先ライン API ---------------------------------------------------------

@app.get("/api/outbound-lines")
def list_outbound_lines():
    """登録された出庫先ラインと、その品目（並び順）を返す。"""
    with get_conn() as conn:
        lines = [
            dict(r)
            for r in conn.execute(
                "SELECT id, name FROM outbound_lines ORDER BY id"
            )
        ]
        items = conn.execute(
            "SELECT oli.line_id, oli.material_id, oli.position, "
            "m.code, m.name, m.product_name, m.pack_size, m.unit, m.active "
            "FROM outbound_line_items oli "
            "JOIN materials m ON m.id = oli.material_id "
            "ORDER BY oli.line_id, oli.position, oli.id"
        ).fetchall()
        by_line: dict[int, list] = {}
        for it in items:
            by_line.setdefault(it["line_id"], []).append({
                "material_id": it["material_id"],
                "code": it["code"],
                "name": it["name"],
                "product_name": it["product_name"],
                "pack_size": it["pack_size"],
                "unit": it["unit"],
                "active": bool(it["active"]),
            })
        for ln in lines:
            ln["items"] = by_line.get(ln["id"], [])
        return lines


@app.post("/api/outbound-lines", status_code=201)
def create_outbound_line(ln: OutboundLineIn):
    with get_conn() as conn:
        try:
            cur = conn.execute(
                "INSERT INTO outbound_lines(name) VALUES (?)", (ln.name,)
            )
        except sqlite3.IntegrityError:
            raise HTTPException(400, f"「{ln.name}」はすでに登録されています")
        return {"id": cur.lastrowid, "name": ln.name, "items": []}


@app.put("/api/outbound-lines/{line_id}")
def update_outbound_line(line_id: int, ln: OutboundLineUpdate):
    """ライン名と品目の並びをまとめて差し替える。"""
    with get_conn() as conn:
        if conn.execute(
            "SELECT 1 FROM outbound_lines WHERE id = ?", (line_id,)
        ).fetchone() is None:
            raise HTTPException(404, "出庫先ラインが見つかりません")
        try:
            conn.execute(
                "UPDATE outbound_lines SET name = ? WHERE id = ?",
                (ln.name, line_id),
            )
        except sqlite3.IntegrityError:
            raise HTTPException(400, f"「{ln.name}」はすでに登録されています")
        conn.execute(
            "DELETE FROM outbound_line_items WHERE line_id = ?", (line_id,)
        )
        seen: set[int] = set()
        for pos, mid in enumerate(ln.items):
            if mid in seen:
                continue
            seen.add(mid)
            if conn.execute(
                "SELECT 1 FROM materials WHERE id = ?", (mid,)
            ).fetchone() is None:
                continue
            conn.execute(
                "INSERT INTO outbound_line_items(line_id, material_id, position) "
                "VALUES (?, ?, ?)",
                (line_id, mid, pos),
            )
    return {"ok": True}


@app.delete("/api/outbound-lines/{line_id}")
def delete_outbound_line(line_id: int):
    with get_conn() as conn:
        cur = conn.execute(
            "DELETE FROM outbound_lines WHERE id = ?", (line_id,)
        )
        if cur.rowcount == 0:
            raise HTTPException(404, "出庫先ラインが見つかりません")
    return {"ok": True}


# --- 入出庫履歴 API -----------------------------------------------------------

@app.get("/api/transactions")
def list_transactions(material_id: Optional[int] = None, limit: int = 200):
    limit = max(1, min(limit, 1000))
    sql = (
        "SELECT t.*, m.code AS material_code, m.name AS material_name, "
        "m.product_name AS material_product, m.unit AS unit "
        "FROM transactions t JOIN materials m ON m.id = t.material_id "
    )
    params: list = []
    if material_id is not None:
        sql += "WHERE t.material_id = ? "
        params.append(material_id)
    sql += "ORDER BY t.created_at DESC, t.id DESC LIMIT ?"
    params.append(limit)
    with get_conn() as conn:
        rows = conn.execute(sql, params).fetchall()
    return [dict(r) for r in rows]


@app.post("/api/transactions", status_code=201)
def create_transaction(tx: TransactionIn):
    with get_conn() as conn:
        if conn.execute("SELECT 1 FROM materials WHERE id = ?", (tx.material_id,)).fetchone() is None:
            raise HTTPException(404, "原料が見つかりません")
        cur = conn.execute(
            "INSERT INTO transactions (material_id, type, quantity, line, note) "
            "VALUES (?, ?, ?, ?, ?)",
            (tx.material_id, tx.type, tx.quantity, tx.line, tx.note),
        )
        if _stock(conn, tx.material_id) < 0:
            raise HTTPException(400, "現在庫を超える出庫は登録できません。数量を確認してください")
        return {"id": cur.lastrowid}


@app.put("/api/transactions/{tx_id}")
def update_transaction(tx_id: int, tx: TransactionIn):
    with get_conn() as conn:
        old = conn.execute("SELECT * FROM transactions WHERE id = ?", (tx_id,)).fetchone()
        if old is None:
            raise HTTPException(404, "履歴が見つかりません")
        if conn.execute("SELECT 1 FROM materials WHERE id = ?", (tx.material_id,)).fetchone() is None:
            raise HTTPException(404, "原料が見つかりません")
        conn.execute(
            "UPDATE transactions SET material_id = ?, type = ?, quantity = ?, line = ?, "
            "note = ? WHERE id = ?",
            (tx.material_id, tx.type, tx.quantity, tx.line, tx.note, tx_id),
        )
        for mid in {old["material_id"], tx.material_id}:
            if _stock(conn, mid) < 0:
                raise HTTPException(400, "この修正を行うと在庫がマイナスになります。内容を確認してください")
    return {"ok": True}


@app.delete("/api/transactions/{tx_id}")
def delete_transaction(tx_id: int):
    with get_conn() as conn:
        old = conn.execute("SELECT * FROM transactions WHERE id = ?", (tx_id,)).fetchone()
        if old is None:
            raise HTTPException(404, "履歴が見つかりません")
        conn.execute("DELETE FROM transactions WHERE id = ?", (tx_id,))
        if _stock(conn, old["material_id"]) < 0:
            raise HTTPException(400, "この記録を削除すると在庫がマイナスになります")
    return {"ok": True}


@app.post("/api/transactions/batch-out")
def create_batch_out(batch: BatchOutIn):
    """出庫依頼表のように複数原料をまとめて出庫登録する。

    在庫が足りる明細だけ登録し、不足している明細は登録せず skipped で返す。
    同じ原料が複数行にある場合は登録済み分を反映した在庫で順に判定する。
    """
    if not batch.items:
        raise HTTPException(400, "出庫する明細がありません")
    with get_conn() as conn:
        registered = 0
        skipped = []
        for item in batch.items:
            m = conn.execute(
                "SELECT name, unit FROM materials WHERE id = ?", (item.material_id,)
            ).fetchone()
            if m is None:
                raise HTTPException(404, "原料が見つかりません")
            available = _stock(conn, item.material_id)
            if available < item.quantity:
                skipped.append(
                    {
                        "material_id": item.material_id,
                        "name": m["name"],
                        "unit": m["unit"],
                        "requested": item.quantity,
                        "available": available,
                    }
                )
                continue
            conn.execute(
                "INSERT INTO transactions (material_id, type, quantity, line, note) "
                "VALUES (?, 'out', ?, ?, ?)",
                (item.material_id, item.quantity, batch.line, batch.note),
            )
            registered += 1
        return {"registered": registered, "skipped": skipped}


# --- Excel 取り込み API -------------------------------------------------------

@app.post("/api/import/materials")
async def import_materials(file: UploadFile = File(...)):
    """原料一覧の Excel(.xlsx / .xls) を読み込み、カタログへ取り込む。

    見出し行（「ｺｰﾄﾞ」を含む行）を自動で探し、必要な列だけ取り込む。
    新規コードはカタログ（active=0）として追加し、既存コードは
    一般名称・原材料名・入数・仕入先のみ更新する（保管場所・発注点・登録状態は維持）。
    """
    content = await file.read()
    rows = _read_workbook(file.filename, content)

    header = []
    header_idx = None
    for i, row in enumerate(rows[:15]):
        cells = [_norm(_cell_str(c)) for c in row]
        if "ｺｰﾄﾞ" in cells or "コード" in cells:
            header, header_idx = cells, i
            break
    if header_idx is None:
        raise HTTPException(400, "見出し行（「ｺｰﾄﾞ」の列）が見つかりませんでした")

    def find(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None

    c_code = find("ｺｰﾄﾞ", "コード")
    c_name = find("一般名称")
    c_product = find("原材料名")
    c_pack = find("入数", "入り数")
    c_supplier = find("購入元1", "購入元")
    if c_code is None:
        raise HTTPException(400, "「ｺｰﾄﾞ」の列が見つかりませんでした")

    def cell(row, idx):
        if idx is None or idx >= len(row):
            return ""
        return _cell_str(row[idx])

    imported = 0
    updated = 0
    with get_conn() as conn:
        existing = {r["code"] for r in conn.execute("SELECT code FROM materials")}
        for row in rows[header_idx + 1:]:
            code = cell(row, c_code)
            if not code:
                continue
            name = cell(row, c_name)
            product = cell(row, c_product)
            pack = cell(row, c_pack)
            supplier = cell(row, c_supplier)
            if code in existing:
                # 既存はExcel由来の項目だけ更新（保管場所・発注点・登録状態は維持）
                conn.execute(
                    "UPDATE materials SET name = ?, product_name = ?, pack_size = ?, "
                    "supplier = ? WHERE code = ?",
                    (name, product, pack, supplier, code),
                )
                updated += 1
            else:
                # 新規はカタログとして取り込む（active=0、在庫登録は別途）
                conn.execute(
                    "INSERT INTO materials "
                    "(code, name, product_name, pack_size, unit, reorder_point, "
                    "supplier, location, active) "
                    "VALUES (?, ?, ?, ?, '個', 0, ?, '倉庫', 0)",
                    (code, name, product, pack, supplier),
                )
                existing.add(code)
                imported += 1
    return {"imported": imported, "updated": updated}


@app.post("/api/import/transactions")
async def import_transactions(file: UploadFile = File(...)):
    """入庫記録の Excel(.xlsx / .xls) を読み込み、入庫履歴へ取り込む（原料用・包材用に対応）。

    商品コードで原料を特定し、見つからなければ品名で照合する。
    すでに倉庫に在庫登録されている原料の行だけを取り込み、未登録の行はスキップする。
    個数を入庫数量、納品日を入庫日として記録する。納品日が未来の行は
    入荷予定として記録され、現在庫には加算されない（納品日が来れば自動で加算）。
    各行の指紋を import_key に保存し、再取り込み時の二重登録を防ぐ。
    """
    content = await file.read()
    rows = _read_workbook(file.filename, content)

    header = []
    header_idx = None
    for i, row in enumerate(rows[:15]):
        cells = [_norm(_cell_str(c)) for c in row]
        if "Fr商品コード" in cells or "商品コード" in cells:
            header, header_idx = cells, i
            break
    if header_idx is None:
        raise HTTPException(
            400, "見出し行（「Fr商品コード」または「商品コード」の列）が見つかりませんでした"
        )

    def find(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None

    c_code = find("Fr商品コード", "商品コード")
    c_qty = find("個数")
    c_date = find("納品日")
    c_maker = find("仕入メーカー", "請求先")
    c_pack = find("単位")
    c_order = find("発注日")
    c_other = find("相手商品コード")
    c_name = find("品名")
    if c_code is None or c_qty is None:
        raise HTTPException(
            400, "「Fr商品コード（商品コード）」または「個数」の列が見つかりませんでした"
        )

    def cell(row, idx):
        if idx is None or idx >= len(row):
            return ""
        return _cell_str(row[idx])

    def raw(row, idx):
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    imported = 0
    planned = 0
    skipped_dup = 0
    skipped = 0
    today = datetime.date.today().strftime("%Y-%m-%d")
    seen: dict[str, int] = {}
    with get_conn() as conn:
        mats_by_code = {}
        mats_by_name = {}
        for r in conn.execute(
            "SELECT id, code, name, product_name, active FROM materials"
        ):
            m = dict(r)
            mats_by_code[m["code"]] = m
            for nm in (m["product_name"], m["name"], m["name"] + m["product_name"]):
                k = _norm(nm)
                if k:
                    mats_by_name.setdefault(k, m)
        existing_keys = {
            r["import_key"]
            for r in conn.execute(
                "SELECT import_key FROM transactions WHERE import_key != ''"
            )
        }
        for row in rows[header_idx + 1:]:
            code = cell(row, c_code)
            pname = cell(row, c_name)
            try:
                qty = float(cell(row, c_qty))
            except ValueError:
                qty = 0
            if qty <= 0 or (not code and not pname):
                continue
            # コードで照合 → 見つからなければ品名で照合
            mat = mats_by_code.get(code) if code else None
            if mat is None and pname:
                mat = mats_by_name.get(_norm(pname))
            # 倉庫に登録済みの原料だけ取り込む。未登録・不明な行はそのままスキップ
            if mat is None or not mat["active"]:
                skipped += 1
                continue
            base = _row_fingerprint([
                code, cell(row, c_order), cell(row, c_date), cell(row, c_qty),
                cell(row, c_other), cell(row, c_maker), pname,
                cell(row, c_pack),
            ])
            seen[base] = seen.get(base, 0) + 1
            key = f"{base}/{seen[base]}"
            if key in existing_keys:
                skipped_dup += 1
                continue
            note = " ".join(
                p for p in [cell(row, c_maker), cell(row, c_pack)] if p
            )
            date_str = _parse_inbound_date(raw(row, c_date))
            conn.execute(
                "INSERT INTO transactions "
                "(material_id, type, quantity, line, note, import_key, created_at) "
                "VALUES (?, 'in', ?, '', ?, ?, ?)",
                (mat["id"], qty, note, key, date_str),
            )
            existing_keys.add(key)
            # 納品日が未来なら入荷予定（現在庫には加算されない）
            if date_str > today:
                planned += 1
            else:
                imported += 1
    return {
        "imported": imported,
        "planned": planned,
        "skipped_dup": skipped_dup,
        "skipped": skipped,
    }


# --- Excel 出力 API -----------------------------------------------------------

@app.get("/api/export/stocktake")
def export_stocktake():
    """棚卸用テンプレートを Excel で返す。数量・メモ列は空欄のまま。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT m.name, m.product_name, m.pack_size, m.unit, m.supplier, m.location,
                   COALESCE((SELECT SUM(CASE
                       WHEN type = 'in'
                            AND date(created_at) <= date('now', 'localtime')
                            THEN quantity
                       WHEN type = 'out' THEN -quantity ELSE 0 END)
                     FROM transactions t WHERE t.material_id = m.id), 0) AS stock
            FROM materials m
            WHERE m.active = 1
            ORDER BY m.location, m.sort_order, m.code
            """
        ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "棚卸"

    today = datetime.date.today()
    headers = ["仕入先", "名称", "入り数", "アプリ数量", "実数", "メモ"]
    widths = [20, 40, 14, 12, 12, 24]

    title = ws.cell(row=1, column=1, value=f"棚卸シート（{today.strftime('%Y/%m/%d')} 出力）")
    title.font = Font(size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 22

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    loc_font = Font(bold=True, color="FFFFFF")
    loc_fill = PatternFill("solid", fgColor="1F4E79")

    r = 4
    cur_loc = None
    for row in rows:
        loc = row["location"] or "倉庫"
        if loc != cur_loc:
            # 保管場所の見出し行
            cell = ws.cell(row=r, column=1, value=loc)
            cell.font = loc_font
            cell.fill = loc_fill
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            for ci in range(2, len(headers) + 1):
                ws.cell(row=r, column=ci).fill = loc_fill
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
            ws.row_dimensions[r].height = 20
            r += 1
            cur_loc = loc
        ws.cell(row=r, column=1, value=row["supplier"] or "")
        ws.cell(row=r, column=2, value=row["product_name"] or row["name"] or "")
        ws.cell(row=r, column=3, value=row["pack_size"] or "")
        ws.cell(row=r, column=4, value=row["stock"])
        # 実数(5)、メモ(6)は空欄のまま
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="right")
        ws.cell(row=r, column=5).alignment = Alignment(horizontal="right")
        for ci in range(1, len(headers) + 1):
            ws.cell(row=r, column=ci).border = border
        ws.row_dimensions[r].height = 24
        r += 1

    ws.freeze_panes = "A4"
    ws.print_title_rows = "3:3"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"stocktake_{today.strftime('%Y%m%d')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# --- 画面 ---------------------------------------------------------------------

@app.get("/")
def index():
    return FileResponse(os.path.join(STATIC_DIR, "index.html"))


if __name__ == "__main__":
    import uvicorn

    host_name = socket.gethostname()
    print("=" * 56)
    print("  倉庫在庫管理アプリ を起動します")
    print(f"  このサーバー上:  http://localhost:{PORT}/")
    print(f"  社内の各 PC から: http://{host_name}:{PORT}/")
    print("  停止するには Ctrl+C を押してください")
    print("=" * 56)
    uvicorn.run(app, host="0.0.0.0", port=PORT)
