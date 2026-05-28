"""初期データ投入スクリプト。

カタログ用 Excel（原材料データベース）と倉庫在庫表（Book1.xlsx 形式）を
受け取り、inventory.db にカタログを取り込んだうえで、在庫表に載っている
原料を倉庫として登録する。在庫表の出現順を sort_order に反映する。

    python seed.py <catalog.xls> <inventory_list.xlsx> [--db inventory.db] [--reset]

--reset を付けると既存 DB を消してから作り直す。
"""

import argparse
import os
import re
import sys
import unicodedata

import openpyxl
from fastapi.testclient import TestClient


_WS = re.compile(r"[\s　]+")
_PAREN_FULL = re.compile(r"[（(][^）)]*[）)]")
_PAREN_TAIL = re.compile(r"[（(].*$")
_SIZE_TAIL = re.compile(
    r"(\d+(?:\.\d+)?)(kg|ｋｇ|㎏|g|ｇ|l|ℓ|ml|ｍｌ|k|本|缶|個|枚|箱|袋|入|ケース|ｹｰｽ|本入|枚入)+缶?$",
    re.IGNORECASE,
)


def _forms(s) -> list[list[str]]:
    """品名から照合用の正規化候補をレベル別に返す（先頭ほど厳密）。"""
    if s is None:
        return []
    raw = unicodedata.normalize("NFKC", str(s)).strip()
    if not raw:
        return []

    def ws(x: str) -> str:
        return _WS.sub("", x).strip()

    raw_ws = ws(raw)
    open_ws = ws(raw.replace("（", "").replace("）", "").replace("(", "").replace(")", ""))
    drop_ws = ws(_PAREN_TAIL.sub("", _PAREN_FULL.sub("", raw)))

    def nosize(x: str) -> str:
        return _SIZE_TAIL.sub("", x)

    def nodash(x: str) -> str:
        return re.sub(r"[\-‐-―・]", "", x)

    levels: list[list[str]] = [
        [raw_ws],
        [open_ws, drop_ws],
        [nosize(open_ws), nosize(drop_ws)],
        [nodash(nosize(open_ws)), nodash(nosize(drop_ws))],
    ]
    # レベルごとに重複と空文字を除去
    out: list[list[str]] = []
    for lv in levels:
        seen: set[str] = set()
        items: list[str] = []
        for x in lv:
            if x and x not in seen:
                seen.add(x)
                items.append(x)
        out.append(items)
    return out


def _read_stock_list(path: str):
    """在庫一覧 Excel を読み、(品名, 商品コード) の並びを返す。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows[:20]):
        if row and any("品名" in str(c) for c in row if c):
            header_idx = i
            break
    if header_idx is None:
        raise SystemExit("見出し行（『品名』を含む行）が見つかりません")
    header = [str(c) if c else "" for c in rows[header_idx]]

    def find_col(*names):
        for n in names:
            for i, h in enumerate(header):
                if n in h:
                    return i
        return None

    c_name = find_col("品名")
    c_code = find_col("商品コード", "コード", "ｺｰﾄﾞ")
    items = []
    for row in rows[header_idx + 1:]:
        if not row or len(row) <= c_name:
            continue
        name = row[c_name]
        if not name or not str(name).strip():
            continue
        code = row[c_code] if c_code is not None and len(row) > c_code else None
        items.append((str(name).strip(), str(code).strip() if code else ""))
    return items


def _code_str(v) -> str:
    if v is None or v == "":
        return ""
    try:
        f = float(v)
        if f.is_integer():
            return str(int(f))
    except (TypeError, ValueError):
        pass
    return str(v).strip()


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("catalog", help="カタログ用 Excel（.xls / .xlsx）")
    p.add_argument("stock_list", help="倉庫在庫表（.xlsx）")
    p.add_argument("--db", default="inventory.db", help="出力する DB ファイル")
    p.add_argument("--reset", action="store_true", help="既存 DB を消してから作り直す")
    args = p.parse_args()

    if args.reset and os.path.exists(args.db):
        os.remove(args.db)

    # database.DB_PATH を切り替えてから main を import
    import database
    database.DB_PATH = os.path.abspath(args.db)
    database.init_db()
    from main import app

    client = TestClient(app)

    # 1) カタログを取り込み
    with open(args.catalog, "rb") as f:
        r = client.post(
            "/api/import/materials",
            files={"file": (os.path.basename(args.catalog), f)},
        )
    if r.status_code != 200:
        print("カタログ取り込み失敗:", r.status_code, r.text)
        return 1
    print(f"カタログ取り込み: {r.json()}")

    # 2) カタログを取得して索引を作る（正規化形 → material、厳密形が優先）
    catalog = client.get("/api/catalog").json()
    by_code = {m["code"]: m for m in catalog}
    idx: dict[str, dict] = {}
    for m in catalog:
        for field in ("product_name", "name"):
            for lv_forms in _forms(m.get(field) or ""):
                for key in lv_forms:
                    idx.setdefault(key, m)
    print(f"カタログ件数: {len(catalog)}")

    # 3) 在庫一覧を読んで一致するものを順番に集める
    stock = _read_stock_list(args.stock_list)
    print(f"在庫一覧の品目: {len(stock)} 件")

    ids: list[int] = []
    seen: set[int] = set()
    unmatched: list[str] = []
    for raw_name, raw_code in stock:
        m = None
        # 商品コード一致を最優先
        code = _code_str(raw_code)
        if code and code in by_code:
            m = by_code[code]
        if m is None:
            # 名称を複数の正規化形で照合（厳密マッチ→曖昧マッチ）
            for lv_forms in _forms(raw_name):
                if m is not None:
                    break
                for key in lv_forms:
                    cand = idx.get(key)
                    if cand is not None:
                        m = cand
                        break
        if m is None:
            unmatched.append(raw_name)
            continue
        if m["id"] in seen:
            continue
        seen.add(m["id"])
        ids.append(m["id"])

    print(f"一致: {len(ids)} 件 / 不一致: {len(unmatched)} 件")
    if unmatched[:8]:
        print("不一致サンプル(先頭8件):")
        for n in unmatched[:8]:
            print(f"  - {n}")

    # 4) 一致したものを倉庫に登録（順番を維持）
    if ids:
        r = client.post(
            "/api/materials/register",
            json={"location": "倉庫", "ids": ids},
        )
        if r.status_code != 200:
            print("在庫登録失敗:", r.status_code, r.text)
            return 1
        print(f"在庫登録: {r.json()}")

    # 5) 結果を確認
    mats = client.get("/api/materials").json()
    print(f"在庫一覧 件数: {len(mats)}")
    print("先頭5件（並び順確認）:")
    for m in mats[:5]:
        print(f"  {m['code']}  {m['name']}  {m['product_name']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
